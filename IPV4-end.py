import os
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import sqlite3
from tkinter import Text
from ttkthemes import ThemedStyle
from tkinter import ttk
import socket
import os
import psutil
import ctypes
import subprocess
import speedtest
from getmac import get_mac_address
import geocoder
import platform
import nmap
from ping3 import ping, verbose_ping
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import requests

# Função para salvar informações em um arquivo de texto


def salvar_informacoes_em_txt(informacoes, arquivo):
    with open(arquivo, 'w') as file:
        file.write(informacoes)

# Função para salvar informações em uma planilha Excel


def salvar_informacoes_em_excel(informacoes, arquivo):
    # Crie um arquivo Excel
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Organize os dados em colunas
    colunas = ["Classe de IP", "Classe de Rede", "Status",
               "Máscara de Sub-rede", "ID de Rede", "Wild Card", "Broadcast"]
    for i, coluna in enumerate(colunas, 1):
        sheet.cell(row=1, column=i, value=coluna)

    info = informacoes.split('\n')
    for i, linha in enumerate(info, 2):
        sheet.cell(row=i, column=1, value=linha)

    # Salve o arquivo Excel
    wb.save(arquivo)


def salvar_e_mostrar_informacoes():
    ip = ip_entry.get()

    if not ip:
        messagebox.showerror(
            "Erro", "Por favor, insira um endereço IP válido.")
        return

    # Resto do código para coletar informações do IP

    info_text.delete(1.0, tk.END)  # Limpa o widget de texto

    def Classe_IP(ip):
        if int(ip.split('.')[0]) < 128:
            return "Classe A"
        elif int(ip.split('.')[0]) < 192:
            return "Classe B"
        else:
            return "Classe C"

    def Classe_Rede(ip):
        ip_parts = ip.split('/')
        prefixo = int(ip_parts[-1])

        if prefixo < 16:
            classe_rede = "Classe A"
        elif prefixo < 24:
            classe_rede = "Classe B"
        else:
            classe_rede = "Classe C"

        return classe_rede
# sub redes são os numros de 1 a 32 exeto os numeros  8, 16, 32

    def classe_sub_rede(ip):
        ip_parts = ip.split('/')
        prefixo = int(ip_parts[-1])
        if prefixo == 16 or prefixo == 24 or prefixo == 8:
            status = "Não é Uma Sub Rede"
        else:
            status = "É uma Sub Rede"
        return status

    def calculo_mascara(ip):
        ip_parts = ip.split('/')
        prefixo = int(ip_parts[-1])
# pega o prefixo apos a barraa e subtrai de acordo  com   o tamnho seguindo  a lei das classes ex:/19
        if prefixo < 16:
            calc = prefixo - 8
        elif prefixo <= 24:
            calc = prefixo - 16
        else:
            calc = prefixo - 32
# segindo com o ex: /19 pertence a classe B da sub rede então  faremos assim 19 - 16 sobrando 3 a quantidade de bits que iremos somar de acordo com a  tabela 1 - 128 2 - 64 ...
        if calc == 1:
            CM = 2**7
        elif calc == 2:
            CM = 2**7 + 2**6
        elif calc == 3:
            CM = 2**7 + 2**6 + 2**5
        elif calc == 4:
            CM = 2**7 + 2**6 + 2**5 + 2**4
        elif calc == 5:
            CM = 2**7 + 2**6 + 2**5 + 2**4 + 2**3
        elif calc == 6:
            CM = 2**7 + 2**6 + 2**5 + 2**4 + 2**3 + 2**2
        elif calc == 7:
            CM = 2**7 + 2**6 + 2**5 + 2**4 + 2**3 + 2**2 + 2**1
        else:
            CM = "erro"  # Definindo um valor padrão para o caso de calc ser inválido
# aki definimos qual o tipo de classe sera a nossa maskara no caso qual layout vamos usar de acordo com a CLasse da Rede
        if Classe_Rede(ip) == "Classe A":
            mascara_rede = f"255.{CM}.0.0"
        elif Classe_Rede(ip) == "Classe B":
            mascara_rede = f"255.255.{CM}.0"
        elif Classe_Rede(ip) == "Classe C":
            mascara_rede = f"255.255.255.{CM}"
        else:
            mascara_rede = "erro"  # Definindo um valor padrão para o caso de classe de rede inválida

        return mascara_rede
# aki pegaremos o id da rede que seria a primeira parte do ip fora o prefixo ex:/19

    def Id_da_rede(ip):
        partes = ip.split('/')
        parte_ip = partes[0]
        return parte_ip
# para acharmos o wild card é nessesario subtrair 255.255.255.255 - mascara de rede que achamos  anteriormente.

    def wild_card():
        mascara = calculo_mascara(ip)
        partes_mascara = mascara.split('.')
        wildcard = []

        for parte in partes_mascara:
            octeto_wildcard = 255 - int(parte)
            wildcard.append(str(octeto_wildcard))

        wildcard_str = ".".join(wildcard)
        return wildcard_str
# ao somar o id da rede com o wildcard temos o broadcast

    def broad_cast():
        wildcard_parts = wild_card().split('.')
        id_parts = Id_da_rede(ip).split('.')

        broadcast_parts = []

        for w, id in zip(wildcard_parts, id_parts):
            broadcast_part = str(int(w) | int(id))
            broadcast_parts.append(broadcast_part)

        broadcast = ".".join(broadcast_parts)
        return broadcast

    data = f"Classe de IP: {Classe_IP(ip)}\n"
    data += f"Classe de Rede: {Classe_Rede(ip)}\n"
    data += f"Status: {classe_sub_rede(ip)}\n"
    data += f"Máscara de Sub-rede: {calculo_mascara(ip)}\n"
    data += f"ID de Rede: {Id_da_rede(ip)}\n"
    data += f"Wild Card: {wild_card()}\n"
    data += f"Broadcast: {broad_cast()}\n"

    # Corrija o problema da pasta de salvamento
    pasta_salvamento = os.path.join(
        os.path.expanduser("~"), "Downloads", "informacoes_ips")
    os.makedirs(pasta_salvamento, exist_ok=True)

    # Caminhos para os arquivos
    txt_file = os.path.join(pasta_salvamento, "informacoes_ip.txt")
    excel_file = os.path.join(pasta_salvamento, "informacoes_ip.xlsx")

    # Salve as informações em um arquivo de texto
    salvar_informacoes_em_txt(data, txt_file)

    # Salve as informações em uma planilha Excel
    salvar_informacoes_em_excel(data, excel_file)

    if data:
        info_text.delete(1.0, tk.END)  # Limpa o widget de texto
        info_text.insert(tk.END, data)
    else:
        messagebox.showerror(
            "Erro", "As informações não foram encontradas no banco de dados.")


def obter_gateway_info():
    gateways = psutil.net_if_addrs()
    gateway_info = "Não foi possível obter informações do roteador"

    for interface, addrs in gateways.items():
        if "Wi-Fi" in interface or "Wireless" in interface:
            connection_type = "Conectado via Wi-Fi"
        elif "Ethernet" in interface:
            connection_type = "Conectado via Ethernet"
        for addr in addrs:
            if addr.family == psutil.AF_LINK:
                gateway_info = f"Endereço MAC do Roteador: {addr.address}"

    return connection_type, gateway_info

# Função para obter a localização com base no IP


def obter_ip_location(ip_address):
    try:
        url = f"http://ipinfo.io/{ip_address}/json"
        response = requests.get(url)
        data = response.json()
        if 'city' in data and 'region' in data and 'country' in data:
            return f"Localização: {data['country']} - {data['region']} - {data['city']}\nLatitude: {data['loc'].split(',')[0]}, Longitude: {data['loc'].split(',')[1]}"
        else:
            return "Localização não encontrada"
    except Exception as e:
        return f"Erro ao obter localização: {str(e)}"

def obter_informacoes_de_rede(info_text2):
    host_name = socket.gethostname()
    ip_address = socket.gethostbyname(host_name)
    system_info = f"Nome do Host: {host_name}\nEndereço IP Local: {ip_address}\nSistema Operacional: {platform.system()} {platform.release()}"

    # Obtenha o nome da rede Wi-Fi
    wifi_name = "Não conectado à Wi-Fi"

    # Obtenha informações da interface de rede
    network_info = (
        f"Informações de Rede:\n"
        f"{system_info}\n"
        f"{wifi_name}\n"
    )

    network_interfaces = psutil.net_if_addrs()
    network_info += "\nInformações da Interface de Rede:\n"
    for interface, addrs in network_interfaces.items():
        network_info += f"Interface: {interface}\n"
        for addr in addrs:
            if addr.family == socket.AF_INET:
                network_info += f"  Endereço IP: {addr.address}\n"
                network_info += f"  Máscara de Sub-rede: {addr.netmask}\n"
            if addr.family == psutil.AF_LINK:
                network_info += f"  Endereço MAC: {addr.address}\n"
        network_info += "\n"

    # Obtenha informações de localização com base no IP
    location_info = obter_ip_location(ip_address)

    # Atualize o widget de texto com as informações
    info_text2.delete('1.0', tk.END)  # Limpe o texto anterior, se houver
    network_info += f"{location_info}\n"
    info_text2.insert('1.0', network_info)

# Crie a janela principal
root = tk.Tk()
root.title("Minhas Definições IP")

# Configure a janela para tela cheia
largura_tela = root.winfo_screenwidth()
altura_tela = root.winfo_screenheight()
root.geometry(f"{largura_tela}x{altura_tela}")
root.attributes('-fullscreen', True)

# Personalize o estilo da janela
style = ttk.Style()
# Use um tema diferente da ttk para melhor integração com o Tkinter
style.theme_use("clam")

# Crie um frame para agrupar os elementos
frame = tk.Frame(root)
frame.pack(fill=tk.BOTH, expand=True)

# Configure o fundo com a cor branca
frame.configure(bg="white")

# Rótulo e campo de entrada para o IP
ip_label = tk.Label(
    frame, text="Digite o endereço IP no formato X.X.X.X/X:", font=('Helvetica', 20))
ip_label.pack(pady=20)

ip_entry = tk.Entry(frame, font=('Helvetica', 18), width=40)
ip_entry.pack(ipadx=20, pady=20)

# Botão para salvar e mostrar informações
process_button = tk.Button(
    frame, text="Salvar e Mostrar", command=salvar_e_mostrar_informacoes, font=('Helvetica', 20), bg='blue', fg='white')
process_button.pack(pady=20)

# Widget de texto para mostrar as informações
info_text = tk.Text(frame, height=15, width=60, font=(
    'Helvetica', 18), bg='black', fg='white')
info_text.pack()

# Crie um segundo frame para as informações de rede
frame2 = tk.Frame(root)
frame2.pack(side=tk.BOTTOM, fill=tk.BOTH)

# Configure o fundo com a cor preta e letras brancas
frame2.configure(bg="black")
style.configure("TLabel", foreground="white")  # Configura o estilo da label

# Widget de texto para mostrar as informações de rede
info_text2 = tk.Text(frame2, height=10, width=largura_tela,
                     font=('Helvetica', 18), bg='black', fg='white')
info_text2.pack()

# Botão para obter informações de rede

# Alinhe o conteúdo ao centro da janela
root.grid_rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)

info_text.insert('1.0', "Insira um endereço IP para exibir informações.")
# Chame a função para exibir informações de rede automaticamente
obter_informacoes_de_rede(info_text2)
root.mainloop()
# Desenvolvido  By: Buehno :)
# contato:  https://linktr.ee/Ronaldo.Bueeno